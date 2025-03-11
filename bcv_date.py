
from datetime import datetime, timedelta
import openpyxl

class ExtraerBCVFecha:
    def __init__(self, fecha):
        self.fecha = fecha
        self.fecha_format_date = datetime.strptime(self.fecha, '%d/%m/%Y')
        
    def ajuste_fecha(self):
        fecha = self.fecha_format_date

        dia = fecha.day
        mes = fecha.month
        # si es 31 poner el 2 y si es 1 tambien el 2
        if (dia == 31 and mes == 12):
            fecha += timedelta(days=2)
        elif (dia == 1 and mes == 1):
            fecha += timedelta(days=1)
        
        #para saltar a lunes
        if fecha.weekday() == 5: #eso significa sabado ya que es el dia 5 
            fecha += timedelta(days=2) 
        elif fecha.weekday() == 6:  #domingo es el dia 6
            fecha += timedelta(days=1) 
        
        return fecha


    def obtener_valor_celda(self):
        archivo = 'recursos/valores.xlsx'
        try:
            #convierte la fecha al ajuste que puse en la func
            fecha = self.ajuste_fecha()
            ano = str(fecha.year)

            self.libro = openpyxl.load_workbook(archivo)

            #por si la hoja  no existe
            if str(ano) not in self.libro.sheetnames:
                return f"La hoja para el {ano} no existe, hay registro desde el 2016."

            #para selccionar la hoja que tiene el ano
            hoja = self.libro[str(ano)]

            indice_fila = 1
            no_encontrado = True

            while no_encontrado:
                fila = hoja.iter_rows(min_row=indice_fila, max_row=indice_fila, max_col=3, values_only=True)
                fila = next(fila)
                if not fila:
                    no_encontrado = False

                
                if type(fila[0]) == str and '/' in  fila[0] :
                    dia = fecha.day
                    mes = fecha.month
                    if len(str(mes)) == 1:
                        mes = f'0{mes}'
                    ano = fecha.year
                    fecha_modificada = f'{dia}/{mes}/{ano}'
                    if fila[0] == fecha_modificada:
                        return fila[2]
                
                if fila[0] == fecha:
                    return fila[2]
                
                indice_fila += 1
            return "Valor no encontrado."
    
        
        except FileNotFoundError:
            print(f"problemas con la data de las tasas del bcv")
            
    def es_hoy(self):     
        return (self.fecha_format_date).date() == (datetime.now()).date()
    
    def ultimo_registro(self):
        hoja = (openpyxl.load_workbook('recursos/valores.xlsx')).active
        valor_ultimo = hoja.cell(row=8, column=1).value

        return str(valor_ultimo).split(' ')[0]
    
    def fecha_existe(self):
        if (self.fecha_format_date > datetime.now()):
            return 'la fecha es del futuro'
        
    def tres_dias(self):
        return abs((self.fecha_format_date.date() - datetime.now().date()).days) <= 3


    def tasaBCV(self):
        """Entrada de la fecha en formato DD/MM/AAAA """
        return str(self.obtener_valor_celda())
    
